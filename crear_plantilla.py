from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Paleta ────────────────────────────────────────────────
NEGRO   = "111111"
BLANCO  = "FFFFFF"
GOLD    = "B8972A"
GOLD_L  = "FEF9C3"
GRAY_H  = "F4F4F5"
GRAY_B  = "E4E4E7"
BLUE_H  = "DBEAFE"
GREEN_H = "DCFCE7"
RED_H   = "FEE2E2"
ORANGE_H= "FFEDD5"

def header_style(cell, dark=True):
    if dark:
        cell.font       = Font(name="Arial", bold=True, color=BLANCO, size=10)
        cell.fill       = PatternFill("solid", fgColor=NEGRO)
    else:
        cell.font       = Font(name="Arial", bold=True, color=NEGRO, size=10)
        cell.fill       = PatternFill("solid", fgColor=GOLD)
    cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border     = thin_border()

def example_style(cell, color=GOLD_L):
    cell.font       = Font(name="Arial", size=9, color="555555", italic=True)
    cell.fill       = PatternFill("solid", fgColor=color)
    cell.alignment  = Alignment(vertical="center")
    cell.border     = thin_border()

def required_style(cell):
    cell.fill       = PatternFill("solid", fgColor=BLUE_H)
    cell.font       = Font(name="Arial", size=10)
    cell.alignment  = Alignment(vertical="center")
    cell.border     = thin_border()

def note_style(cell, color=GRAY_H):
    cell.font       = Font(name="Arial", size=9, color="71717A")
    cell.fill       = PatternFill("solid", fgColor=color)
    cell.alignment  = Alignment(vertical="center", wrap_text=True)
    cell.border     = thin_border()

def thin_border():
    s = Side(style="thin", color=GRAY_B)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, cols_widths):
    for col, width in cols_widths.items():
        ws.column_dimensions[col].width = width

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ══════════════════════════════════════════════════════════
# HOJA 1: INSTRUCCIONES
# ══════════════════════════════════════════════════════════
ws_inst = wb.active
ws_inst.title = "INSTRUCCIONES"
ws_inst.sheet_properties.tabColor = NEGRO

ws_inst.column_dimensions["A"].width = 5
ws_inst.column_dimensions["B"].width = 30
ws_inst.column_dimensions["C"].width = 60
ws_inst.row_dimensions[1].height = 60

# Título
ws_inst.merge_cells("B1:C1")
c = ws_inst["B1"]
c.value = "Culinary · Plantilla de Importación de Datos"
c.font  = Font(name="Georgia", bold=False, size=20, color=NEGRO)
c.fill  = PatternFill("solid", fgColor=GOLD_L)
c.alignment = Alignment(horizontal="center", vertical="center")

pasos = [
    ("", ""),
    ("PASO 1 — ANTES DE COMENZAR", "Descarga este archivo y guárdalo en tu computador."),
    ("", ""),
    ("PASO 2 — LLENA LOS DATOS", "Completa las hojas en este orden:"),
    ("", "  1.  INGREDIENTES  →  precios y proveedores"),
    ("", "  2.  SESIONES      →  una fila por clase"),
    ("", "  3.  DET_SESION    →  ingredientes usados en cada clase"),
    ("", "  4.  INVENTARIO    →  stock actual (opcional)"),
    ("", ""),
    ("PASO 3 — REVISA LAS LISTAS", "La hoja LISTAS_REFERENCIA tiene los valores válidos para:"),
    ("", "  •  Categorías de ingredientes"),
    ("", "  •  Tipos de taller (TC1, TC2, TP1, etc.)"),
    ("", "  •  Nombres de proveedores"),
    ("", ""),
    ("PASO 4 — IMPORTA", "En el sistema Culinary, ve a la sección 'Importar' del menú"),
    ("", "y sube este archivo Excel. El sistema revisará los datos y te"),
    ("", "mostrará un resumen antes de confirmar."),
    ("", ""),
    ("REGLAS IMPORTANTES", ""),
    ("", "✓  No borres ni renombres las columnas con encabezado"),
    ("", "✓  Las celdas en AZUL son OBLIGATORIAS"),
    ("", "✓  Las celdas en AMARILLO son ejemplos — puedes borrarlos"),
    ("", "✓  No modifiques la hoja LISTAS_REFERENCIA"),
    ("", "✓  Guarda el archivo como .xlsx (no .xls ni .csv)"),
    ("", ""),
    ("SOPORTE", "Si tienes dudas, contacta al administrador del sistema."),
]

for i, (titulo, texto) in enumerate(pasos, start=2):
    ws_inst.row_dimensions[i].height = 18
    b = ws_inst.cell(row=i, column=2, value=titulo)
    c2 = ws_inst.cell(row=i, column=3, value=texto)
    if titulo and not titulo.startswith(" ") and titulo != "":
        b.font = Font(name="Arial", bold=True, size=10, color=NEGRO)
        b.fill = PatternFill("solid", fgColor=GRAY_H)
    else:
        b.font = Font(name="Arial", size=10, color="444444")
    c2.font = Font(name="Arial", size=10, color="444444")
    if titulo in ("PASO 1 — ANTES DE COMENZAR","PASO 2 — LLENA LOS DATOS",
                  "PASO 3 — REVISA LAS LISTAS","PASO 4 — IMPORTA"):
        b.fill = PatternFill("solid", fgColor=GOLD)
        b.font = Font(name="Arial", bold=True, size=10, color=BLANCO)


# ══════════════════════════════════════════════════════════
# HOJA 2: INGREDIENTES
# ══════════════════════════════════════════════════════════
ws_ing = wb.create_sheet("INGREDIENTES")
ws_ing.sheet_properties.tabColor = "3B82F6"

headers_ing = [
    ("codigo",         "Código",          "Número del ingrediente (ej: 10040)"),
    ("nombre",         "Nombre *",         "Nombre en MAYÚSCULAS (ej: ACEITE DE OLIVA)"),
    ("precio_neto_clp","Precio Neto CLP *","Precio sin IVA en pesos chilenos (ej: 14516)"),
    ("unidad",         "Unidad *",         "kg / g / L / ml / UN / lt"),
    ("categoria",      "Categoría *",      "Ver hoja LISTAS_REFERENCIA"),
    ("proveedor",      "Proveedor",        "Nombre exacto del proveedor (ver LISTAS_REFERENCIA)"),
]

ws_ing.row_dimensions[1].height = 14
ws_ing.row_dimensions[2].height = 36
ws_ing.row_dimensions[3].height = 18

# Sub-header de instrucción
ws_ing.merge_cells("A1:F1")
c = ws_ing["A1"]
c.value = "✦  Complete una fila por ingrediente  ·  Campos con * son obligatorios  ·  Las filas en amarillo son ejemplos"
c.font  = Font(name="Arial", size=9, color="555555")
c.fill  = PatternFill("solid", fgColor=GOLD_L)
c.alignment = Alignment(horizontal="center", vertical="center")

for col, (field, label, note) in enumerate(headers_ing, start=1):
    c_lbl = ws_ing.cell(row=2, column=col, value=label)
    header_style(c_lbl, dark=True)
    c_note = ws_ing.cell(row=3, column=col, value=note)
    note_style(c_note)

# Ejemplos
ejemplos_ing = [
    ("10040", "ACEITE DE OLIVA",        14516, "L",  "Aceites y Grasas",  "Bidfood Chile SA"),
    ("10011", "HUEVO EXTRA BLANCO",     180,   "UN", "Huevos",            "Caserita"),
    ("10050", "COBERTURA AMARGA 70%",   9800,  "kg", "Chocolates",        "La Vinoteca"),
    ("",      "← Borra estas filas y escribe tus datos", "", "", "", ""),
]

for r, row in enumerate(ejemplos_ing, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws_ing.cell(row=r, column=c, value=val)
        example_style(cell, GOLD_L)

# Filas vacías listas para llenar (azules = obligatorias)
for r in range(8, 108):
    ws_ing.row_dimensions[r].height = 17
    for c in range(1, 7):
        cell = ws_ing.cell(row=r, column=c)
        required_style(cell) if c in (2,3,4,5) else note_style(cell, BLANCO)

set_col_width(ws_ing, {"A":14,"B":32,"C":16,"D":9,"E":20,"F":24})
freeze(ws_ing, "A4")

# Validación unidad
dv_unidad = DataValidation(type="list", formula1='"kg,g,L,ml,UN,lt,pote,bandeja"', allow_blank=True)
ws_ing.add_data_validation(dv_unidad)
dv_unidad.sqref = "D4:D107"


# ══════════════════════════════════════════════════════════
# HOJA 3: SESIONES
# ══════════════════════════════════════════════════════════
ws_ses = wb.create_sheet("SESIONES")
ws_ses.sheet_properties.tabColor = "10B981"

headers_ses = [
    ("cod_clase",     "Código Clase",    "Ej: TCCMSES20 (único por clase)"),
    ("tipo_taller",   "Tipo Taller *",   "Ver LISTAS_REFERENCIA (TC1, TP1, etc.)"),
    ("semana_numero", "N° Semana *",      "Número de semana (1 al 20)"),
    ("anio",          "Año *",            "Ej: 2025"),
    ("fecha",         "Fecha *",          "Formato: YYYY-MM-DD (ej: 2025-04-07)"),
    ("seccion",       "Sección *",        "Número de sección (1, 2, 3...)"),
    ("num_alumnos",   "N° Alumnos *",     "Número de estudiantes en la clase"),
    ("profesor",      "Profesor",         "Nombre del docente (ej: EMERSON O.)"),
]

ws_ses.merge_cells("A1:H1")
c = ws_ses["A1"]
c.value = "✦  Una fila por sesión/clase  ·  El Código Clase debe ser único y coincidir con DET_SESION"
c.font  = Font(name="Arial", size=9, color="555555")
c.fill  = PatternFill("solid", fgColor=GREEN_H)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_ses.row_dimensions[1].height = 14
ws_ses.row_dimensions[2].height = 36
ws_ses.row_dimensions[3].height = 18

for col, (field, label, note) in enumerate(headers_ses, start=1):
    header_style(ws_ses.cell(row=2, column=col, value=label))
    note_style(ws_ses.cell(row=3, column=col, value=note))

ejemplos_ses = [
    ("TCCMSES20", "TC1", 5, 2025, "2025-04-07", 1, 18, "EMERSON O."),
    ("TIPSES25",  "TP1", 5, 2025, "2025-04-08", 2, 15, "FERNANDA M."),
    ("",          "← Borra y completa con tus datos", "", "", "", "", "", ""),
]
for r, row in enumerate(ejemplos_ses, start=4):
    for c, val in enumerate(row, start=1):
        example_style(ws_ses.cell(row=r, column=c, value=val), GREEN_H)

for r in range(7, 107):
    ws_ses.row_dimensions[r].height = 17
    for c in range(1, 9):
        cell = ws_ses.cell(row=r, column=c)
        required_style(cell) if c in (2,3,4,5,6,7) else note_style(cell, BLANCO)

set_col_width(ws_ses, {"A":16,"B":13,"C":11,"D":7,"E":15,"F":9,"G":12,"H":20})
freeze(ws_ses, "A4")


# ══════════════════════════════════════════════════════════
# HOJA 4: DET_SESION
# ══════════════════════════════════════════════════════════
ws_det = wb.create_sheet("DET_SESION")
ws_det.sheet_properties.tabColor = "F97316"

headers_det = [
    ("cod_clase",          "Código Clase *",      "Debe coincidir con hoja SESIONES"),
    ("codigo_ingrediente",  "Código Ingrediente *","Código del ingrediente (hoja INGREDIENTES)"),
    ("nombre_ingrediente",  "Nombre Ingrediente",  "Referencia visual — no se importa"),
    ("cantidad_por_alumno", "Cant. por Alumno *",  "Cantidad que usa cada alumno (ej: 0.05)"),
    ("cantidad_total",      "Cantidad Total *",    "Cant/alumno × N°alumnos (ej: 0.9)"),
    ("unidad",              "Unidad",              "kg / L / UN — debe coincidir con el ingrediente"),
]

ws_det.merge_cells("A1:F1")
c = ws_det["A1"]
c.value = "✦  Una fila por ingrediente utilizado en cada sesión  ·  Una sesión puede tener múltiples filas"
c.font  = Font(name="Arial", size=9, color="555555")
c.fill  = PatternFill("solid", fgColor=ORANGE_H)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_det.row_dimensions[1].height = 14
ws_det.row_dimensions[2].height = 36
ws_det.row_dimensions[3].height = 18

for col, (field, label, note) in enumerate(headers_det, start=1):
    header_style(ws_det.cell(row=2, column=col, value=label))
    note_style(ws_det.cell(row=3, column=col, value=note))

ejemplos_det = [
    ("TCCMSES20", "10040", "ACEITE DE OLIVA",    0.05, 0.9,  "L"),
    ("TCCMSES20", "10011", "HUEVO EXTRA BLANCO", 2,    36,   "UN"),
    ("TCCMSES20", "10050", "COBERTURA AMARGA",   0.1,  1.8,  "kg"),
    ("TIPSES25",  "10040", "ACEITE DE OLIVA",    0.03, 0.45, "L"),
    ("",          "← Agrega un ingrediente por fila", "", "", "", ""),
]
for r, row in enumerate(ejemplos_det, start=4):
    for c, val in enumerate(row, start=1):
        example_style(ws_det.cell(row=r, column=c, value=val), ORANGE_H)

for r in range(9, 309):
    ws_det.row_dimensions[r].height = 17
    for c in range(1, 7):
        cell = ws_det.cell(row=r, column=c)
        required_style(cell) if c in (1,2,4,5) else note_style(cell, BLANCO)

set_col_width(ws_det, {"A":16,"B":18,"C":30,"D":16,"E":15,"F":9})
freeze(ws_det, "A4")


# ══════════════════════════════════════════════════════════
# HOJA 5: INVENTARIO
# ══════════════════════════════════════════════════════════
ws_inv = wb.create_sheet("INVENTARIO")
ws_inv.sheet_properties.tabColor = "8B5CF6"

headers_inv = [
    ("codigo_ingrediente", "Código Ingrediente *","Código del ingrediente"),
    ("nombre_ingrediente", "Nombre Ingrediente",  "Referencia visual"),
    ("cantidad",           "Cantidad en Stock *", "Número con decimales (ej: 3.5)"),
    ("mes",                "Mes *",               "ENERO, FEBRERO, MARZO ... DICIEMBRE"),
    ("anio",               "Año *",               "Ej: 2025"),
]

ws_inv.merge_cells("A1:E1")
c = ws_inv["A1"]
c.value = "✦  Stock actual por ingrediente  ·  Un registro por ingrediente por mes"
c.font  = Font(name="Arial", size=9, color="555555")
c.fill  = PatternFill("solid", fgColor="EDE9FE")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_inv.row_dimensions[1].height = 14
ws_inv.row_dimensions[2].height = 36
ws_inv.row_dimensions[3].height = 18

for col, (field, label, note) in enumerate(headers_inv, start=1):
    header_style(ws_inv.cell(row=2, column=col, value=label))
    note_style(ws_inv.cell(row=3, column=col, value=note))

ejemplos_inv = [
    ("10040", "ACEITE DE OLIVA",       5.5, "ABRIL", 2025),
    ("10011", "HUEVO EXTRA BLANCO",    120,  "ABRIL", 2025),
    ("10050", "COBERTURA AMARGA 70%",  2.0, "ABRIL", 2025),
]
for r, row in enumerate(ejemplos_inv, start=4):
    for c, val in enumerate(row, start=1):
        example_style(ws_inv.cell(row=r, column=c, value=val), "EDE9FE")

dv_mes = DataValidation(type="list",
    formula1='"ENERO,FEBRERO,MARZO,ABRIL,MAYO,JUNIO,JULIO,AGOSTO,SEPTIEMBRE,OCTUBRE,NOVIEMBRE,DICIEMBRE"',
    allow_blank=True)
ws_inv.add_data_validation(dv_mes)
dv_mes.sqref = "D7:D207"

for r in range(7, 207):
    ws_inv.row_dimensions[r].height = 17
    for c in range(1, 6):
        cell = ws_inv.cell(row=r, column=c)
        required_style(cell) if c in (1,3,4,5) else note_style(cell, BLANCO)

set_col_width(ws_inv, {"A":18,"B":30,"C":18,"D":14,"E":8})
freeze(ws_inv, "A4")


# ══════════════════════════════════════════════════════════
# HOJA 6: LISTAS_REFERENCIA
# ══════════════════════════════════════════════════════════
ws_ref = wb.create_sheet("LISTAS_REFERENCIA")
ws_ref.sheet_properties.tabColor = "6B7280"
ws_ref.sheet_view.showGridLines = False

ws_ref.column_dimensions["A"].width = 4
ws_ref.column_dimensions["B"].width = 26
ws_ref.column_dimensions["C"].width = 4
ws_ref.column_dimensions["D"].width = 26
ws_ref.column_dimensions["E"].width = 4
ws_ref.column_dimensions["F"].width = 28

ws_ref.row_dimensions[1].height = 30
ws_ref.merge_cells("B1:F1")
c = ws_ref["B1"]
c.value = "LISTAS DE REFERENCIA — No modificar este contenido"
c.font  = Font(name="Arial", bold=True, size=13, color=BLANCO)
c.fill  = PatternFill("solid", fgColor=NEGRO)
c.alignment = Alignment(horizontal="center", vertical="center")

# Columna B: Categorías
cats = ["Aceites y Grasas","Carnes y Aves","Chocolates","Condimentos",
        "Embutidos","Especias y Hierbas","Frutas y Verduras","Huevos",
        "Lácteos","Mariscos y Pescados","Panadería y Pastelería","Vinos y Licores"]

c = ws_ref["B2"]; c.value = "CATEGORÍAS DE INGREDIENTES"
c.font = Font(name="Arial", bold=True, size=10, color=BLANCO)
c.fill = PatternFill("solid", fgColor=NEGRO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_ref.row_dimensions[2].height = 22

for i, cat in enumerate(cats, start=3):
    cell = ws_ref.cell(row=i, column=2, value=cat)
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=BLUE_H if i%2==0 else BLANCO)
    cell.alignment = Alignment(vertical="center")
    cell.border = thin_border()
    ws_ref.row_dimensions[i].height = 17

# Columna D: Tipos de taller
talleres = [
    ("TC1",  "Técnicas de Cocina 1"),
    ("TC2",  "Técnicas de Cocina 2"),
    ("TP1",  "Técnicas de Pastelería 1"),
    ("TP2",  "Técnicas de Pastelería 2"),
    ("TDLC", "Técnicas de la Cocina Chilena"),
    ("TIC",  "Técnicas de la Int. Culinaria"),
    ("TM",   "Taller de Marketing"),
    ("TPP",  "Taller de Práctica Profesional"),
]

c = ws_ref["D2"]; c.value = "TIPOS DE TALLER (código — nombre)"
c.font = Font(name="Arial", bold=True, size=10, color=BLANCO)
c.fill = PatternFill("solid", fgColor=NEGRO)
c.alignment = Alignment(horizontal="center", vertical="center")

for i, (cod, nom) in enumerate(talleres, start=3):
    cell_c = ws_ref.cell(row=i, column=4, value=f"{cod}  →  {nom}")
    cell_c.font = Font(name="Arial", size=10)
    cell_c.fill = PatternFill("solid", fgColor=GREEN_H if i%2==0 else BLANCO)
    cell_c.alignment = Alignment(vertical="center")
    cell_c.border = thin_border()

# Columna F: Proveedores
provs = ["Bidfood Chile SA","Caserita","Gourmet Select","La Vinoteca",
         "Neucober","Queso Express","Santa Dolores","San Jorge Ltda","Distribuidora FoodPro"]

c = ws_ref["F2"]; c.value = "PROVEEDORES REGISTRADOS"
c.font = Font(name="Arial", bold=True, size=10, color=BLANCO)
c.fill = PatternFill("solid", fgColor=NEGRO)
c.alignment = Alignment(horizontal="center", vertical="center")

for i, prov in enumerate(provs, start=3):
    cell = ws_ref.cell(row=i, column=6, value=prov)
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=ORANGE_H if i%2==0 else BLANCO)
    cell.alignment = Alignment(vertical="center")
    cell.border = thin_border()

# Proteger hoja de referencia
ws_ref.protection.sheet = True
ws_ref.protection.password = "culinary"

# ══════════════════════════════════════════════════════════
# GUARDAR
# ══════════════════════════════════════════════════════════
wb.save(r"C:\Users\dmoov\Documents\Programas\costos-gastronomia\plantilla_importacion.xlsx")
print("OK — plantilla creada")
